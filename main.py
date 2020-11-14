import requests
import logging
import json
import os
from requests import Session
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from openpyxl import Workbook
import openpyxl

logging.basicConfig(filename="all_log.log", filemode='w', level=logging.DEBUG,
                    format='%(asctime)s : %(levelname)s : %(message)s')


class Excel(object):
    def __init__(self, name):
        self.path = os.path.normpath(os.path.dirname(__file__)) + '\\' + name
        self.wb = Workbook()
        self.sheet = self.wb.active

    def read_excel(self):
        try:
            excel_dict = {}
            wb = openpyxl.load_workbook(filename=self.path, data_only=True)
            ws = wb.worksheets[0]
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 1).value and ws.cell(row, 2).value:
                    excel_dict.update({str(ws.cell(row, 1).value): str(ws.cell(row, 2).value)})
            wb.close()
            return excel_dict
        except Exception as error:
            logging.critical(f"exception in read_excel: {error}")

    def write_excel(self, row, bs, data):
        try:
            self.sheet.cell(1, 1).value = 'Old Name'
            self.sheet.cell(1, 2).value = 'New Name'
            self.sheet.cell(row, 1).value = bs
            self.sheet.cell(row, 2).value = data
            self.wb.save(self.path)
            self.wb.close()
        except Exception as error:
            logging.critical(f"exception in write_excel: {error}")


class EnmRestApi(Session):
    def __init__(self, url, login, password):
        super().__init__()
        self.url = url
        self.verify = False
        self.form = {
            'IDToken1': login,
            'IDToken2': password
        }
        requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

    def check_login(self):
        try:
            response = self.post(f'{self.url}/login', data=self.form, timeout=30)
            if response.status_code != requests.codes.ok:
                logging.error(f'incorrect login or password. status_code {response.status_code}')
                print(f'Incorrect Login or Password')
                return False
            else:
                return True
        except Exception as error:
            logging.critical(f"exception in check_login: {error}")
            return False

    def exit(self):
        try:
            self.get(f"{self.url}logout", timeout=30)
        finally:
            super().__exit__(self)

    def get_poid(self, element):
        port_poid = ''
        element_poid = ''
        mo_type = ''
        url = f"https://enm.telecom.com/managedObjects/temporaryQueryForMoClassMapping/v2?query=" \
              f"select%20all%20objects%20of%20type%20NetworkElement%2C%20all%20objects%20of%20type%20" \
              f"ComConnectivityInformation%2C%20all%20objects%20of%20type%20" \
              f"CppConnectivityInformation%20from%20{element}"
        try:
            response = self.get(url, timeout=60)
            resp_dict = response.json()
            logging.debug(resp_dict)

            def _finditem(obj, key):
                if key in obj:
                    return obj[key]
                for k, v in obj.items():
                    if isinstance(v, dict):
                        return _finditem(v, key)

            for moDetail in resp_dict['moDetails']:
                dicts = _finditem(moDetail, 'moTypes')
                com_po_id = _finditem(dicts, 'ComConnectivityInformation')
                cpp_po_id = _finditem(dicts, 'CppConnectivityInformation')
                element_po_id = _finditem(dicts, 'NetworkElement')
                if element_po_id:
                    element_poid = element_po_id[0]['poId']
                if com_po_id:
                    port_poid = com_po_id[0]['poId']
                    mo_type = "ComConnectivityInformation"
                elif cpp_po_id:
                    port_poid = cpp_po_id[0]['poId']
                    mo_type = "CppConnectivityInformation"
            return element_poid, port_poid, mo_type, element
        except Exception as error:
            logging.critical(f"exception in get_poid: {error}")

    def get_pos_by_poids(self, element_poid, port_poid, mo_type, element):
        url = "https://enm.telecom.com/managedObjects/getPosByPoIds"
        try:
            self.headers.update({"Content-Type": "application/json", "X-Requested-With": "XMLHttpRequest"})
            data = {"poList": [element_poid], "defaultMappings": ["syncStatus"], "attributeMappings": [
                {"moType": "NetworkElement",
                 "attributeNames": ["neType", "ossModelIdentity", "ossPrefix", "timeZone", "controllingRnc",
                                    "controllingBsc"]}]}
            response = self.post(url, data=json.dumps(data), timeout=60)
            resp_dict = response.json()
            logging.debug(resp_dict)
            if type(resp_dict) is dict:
                logging.critical(f"exception in get_pos_by_poids element: {element}\n resp_dict: {resp_dict}")
                print(f"Can't get Data BS {element} from ENM. Check BS Name in ENM")
                return False, element
            attributes = resp_dict[0]['attributes']
            data = {"poList": [port_poid], "defaultMappings": ["syncStatus"],
                    "attributeMappings": [{"moType": mo_type, "attributeNames": ["port", "ipAddress"]}]}
            response = self.post(url, data=json.dumps(data), timeout=60)
            resp_dict = response.json()
            logging.debug(resp_dict)
            port = resp_dict[0]['attributes']
            return attributes, port
        except Exception as error:
            logging.critical(f"exception in get_pos_by_poids: {error}")
            print(f"Can't get Data BS {element} from ENM Check BS Name in ENM")
            return False, element

    def cli_app(self, command):
        import binascii
        import os

        def encode_multipart_formdata(fields):
            boundary = binascii.hexlify(os.urandom(8)).decode('ascii')
            body = (
                    "".join("------WebKitFormBoundary%s\r\n"
                            "Content-Disposition: form-data; name=\"%s\"\r\n"
                            "\r\n"
                            "%s\r\n" % (boundary, field, value)
                            for field, value in fields.items()) +
                    "------WebKitFormBoundary%s--" % boundary
            )
            content_type = "multipart/form-data; boundary=----WebKitFormBoundary%s" % boundary
            return body, content_type

        body, content_type = encode_multipart_formdata({"command": command})
        url = "https://enm.telecom.com/script-engine/services/command/"
        try:
            self.headers.update({"Content-Type": content_type,
                                 "X-Requested-With": "XMLHttpRequest", "X-Tor-Application": "cliapp"})
            response = self.post(url, data=body, timeout=60)
            resp_dict = response
            print(resp_dict)
            print(resp_dict.headers)
            print(resp_dict.text)
            process_id = resp_dict.headers['process_id']
            request_id = resp_dict.headers['request_id']
            url = f"https://enm.telecom.com/script-engine/services/command/output/{process_id}?max_size=20000"
            response = self.get(url)
            resp_dict = response
            print(resp_dict)
            print(resp_dict.headers)
            print(resp_dict.text)
            return resp_dict
        except Exception as error:
            logging.critical(f"exception in cli_app: {error}")


class Script(object):
    def __init__(self, old_name, new_name):
        try:
            self.old_name = old_name
            self.new_name = new_name
            self.path = os.path.normpath(os.path.dirname(__file__))
            self.file = open(self.path + '\\' + self.old_name + '.txt', 'w')
            self.all_file = open(self.path + '\\' + 'bs_all.txt', 'a')
        except Exception as error:
            logging.critical(f"exception in __init__ script: {error}")
            print("Can't open or find excel file")

    def write(self, attributes, port, mo_type):
        self.script = f"""cmedit set NetworkElement={self.old_name},PmFunction=1 pmEnabled=false
cmedit set NetworkElement={self.old_name},CmNodeHeartbeatSupervision=1 active=false
cmedit set NetworkElement={self.old_name},InventorySupervision=1 active=false
alarm disable {self.old_name}
cmedit action NetworkElement={self.old_name},CmFunction=1 deleteNrmDataFromEnm
cmedit delete NetworkElement={self.old_name} -ALL --force
cmedit create NetworkElement={self.new_name} networkElementId={self.new_name}, neType={attributes['neType']}, ossModelIdentity="{attributes['ossModelIdentity']}", ossPrefix="{str(attributes['ossPrefix']).split('MeContext')[0]}MeContext={self.new_name}" -ns=OSS_NE_DEF -version=2.0.0
cmedit create NetworkElement={self.new_name},{mo_type}=1 {mo_type}Id=1, ipAddress="{port['ipAddress']}", port={port['port']} -ns=COM_MED -version=1.1.0
secadm credentials create --secureusername rbsuser --secureuserpassword rbsuser1 -n {self.new_name}
cmedit set NetworkElement={self.new_name},CmNodeHeartbeatSupervision=1 active=true
cmedit set NetworkElement={self.new_name},InventorySupervision=1 active=true
cmedit set NetworkElement={self.new_name},PmFunction=1 pmEnabled=true
alarm enable {self.new_name}
cmedit set {self.new_name} NetworkElement timeZone={attributes['timeZone']}
"""
        connect_rnc = f"""cmedit set NetworkElement={self.new_name} controllingRnc="{attributes.get('controllingRnc')}"
"""
        connect_bsc = f"""cmedit set NetworkElement={self.new_name} controllingBsc="{attributes.get('controllingBsc')}"
"""
        if attributes.get('controllingRnc'):
            self.script = self.script + connect_rnc
        if attributes.get('controllingBsc'):
            self.script = self.script + connect_bsc
        try:
            self.file.write(self.script)
            self.file.close()
            self.all_file.write(self.script + '\n')
            self.all_file.close()
            print('Save ' + self.old_name)
        except Exception as error:
            logging.critical(f"exception in write script: {error}")
            print("Can't write or close out txt files")


def main():
    error_bs = 'Error BS: '
    print('Put the file with name bs.xlsx in the current directory with this script and after, '
          'insert your login and password to ENM')
    excel = Excel("bs.xlsx")
    excel_dict = excel.read_excel()
    logging.debug(excel_dict)
    login = input('Insert Login to ENM:')
    password = input('Insert Password to ENM:')
    enm = EnmRestApi('https://enm.telecom.com/', login, password)
    if enm.check_login():
        if os.path.isfile(os.path.normpath(os.path.dirname(__file__)) + '\\' + 'bs_all.txt'):
            os.remove(os.path.normpath(os.path.dirname(__file__)) + '\\' + 'bs_all.txt')
        for old_bs, new_bs in excel_dict.items():
            element_poid, port_poid, mo_type, element = enm.get_poid(old_bs)
            attributes, port = enm.get_pos_by_poids(element_poid, port_poid, mo_type, element)
            if attributes:
                Script(old_bs, new_bs).write(attributes, port, mo_type)
            else:
                error_bs += port + " "
        print('Finished. Check output txt files. ' + error_bs)
        logging.debug('End')
        enm.exit()
        input('Push Any Button to Close Script')
    else:
        return main()


main()

# def test():
#     login = input('Insert Login to Enm:')
#     password = input('Insert Password to Enm:')
#     enm = EnmRestApi('https://enm.telecom.com/', login, password)
#     if enm.check_login():
#         enm.cli_app('cmedit get GL2*')
#         enm.exit()
#
#
# test()
